import openpyxl
import pandas as pd
import numpy as np
import re
import networkx as nx


class LegalText:
    def __init__(self, name, translated=False, references=None):
        self.name = name
        self.translated = translated
        if references is None:
            references = []
        self.references = references

    def __str__(self):
        return f"{self.name}"

    def __repr__(self):
        return self.__str__()


def setup_graph_matrix(index2article_with_references: dict, article2index: dict, option: int = 0) -> np.ndarray:
    if not 0 <= option <= 3:
        print("Invalid value of the option. Please choose an option value between 0, 1, 2 and 3")
        exit(1)
    counter = 0
    graph_matrix = np.zeros((len(index2article_with_references), len(index2article_with_references)), dtype=int)
    for i in range(len(index2article_with_references)):
        # graph_matrix[0][i] = i
        # graph_matrix[i][0] = i
        references = index2article_with_references[i].references
        if references:
            for reference in references:
                regex1 = r"^Art.\ \d+(\(.\))*$"
                regex2 = r"Art\.\ \d+-\d+"
                match1 = re.search(pattern=regex1, string=reference)
                if match1:
                    match1 = match1.string
                    match1 = re.sub(r"[a-zA-Z]+\.\s", "", match1)
                    if option == 2 or option == 3:
                        k = article2index[match1]
                        name = index2article_with_references[k].name
                        while match1 in name:
                            graph_matrix[i][k] = 1
                            counter += 1
                            k += 1
                            if k == len(index2article_with_references):
                                break
                            name = index2article_with_references[k].name
                    else:
                        reference_index = article2index[match1]
                        graph_matrix[i][reference_index] = 1
                        counter += 1
                else:
                    match2 = re.search(pattern=regex2, string=reference)
                    if match2:
                        match2 = match2.string
                        match2 = re.sub(r"[^\d+\-\d+]", "", match2)
                        match2 = match2.split("-")
                        if option == 1 or option == 3:
                            increment = 0
                            interval_start = article2index[match2[0]]
                            interval_end = article2index[match2[1]]
                            k = interval_end
                            name = index2article_with_references[k].name
                            while match2[1] in name:
                                increment += 1
                                k += 1
                                if k == len(index2article_with_references):
                                    break
                                name = index2article_with_references[k].name
                        else:
                            increment = 1
                            interval_start = int(match2[0])
                            interval_end = int(match2[1])
                        for j in range(interval_start, interval_end + increment):
                            if option == 1 or option == 3:
                                reference_index = j
                            else:
                                reference_index = article2index[str(j)]
                            graph_matrix[i][reference_index] = 1
                            counter += 1
    print("Graph matrix has been set up.")
    print(f"Average number of references starting from each article: {counter}/{len(index2article_with_references)} = {counter/len(index2article_with_references)}\n")
    return graph_matrix


def setup_links() -> list:
    wb = openpyxl.load_workbook('GDPR_map.xlsx')
    ws = wb['Cartel1']
    links = []
    for i in range(2, ws.max_row + 1):
        link = ws.cell(row=i, column=6).hyperlink
        if link is None:
            print("No link found.")
        links.append(link.target)
    return links


def setup_dictionaries(df, index2article, index2article_with_references, article2index):
    index2references: dict[int, list] = {}
    for i in range(len(df)):
        row = df.loc[i]
        article, paragraph, point, translated, references, link = row
        if len(references) == 0:
            index2references[i] = []
        else:
            index2references[i] = references.split(', ')
        if len(point) > 0:
            name = article + '(' + paragraph + ')' + '(' + point + ')'
        elif len(paragraph) > 0:
            name = article + '(' + paragraph + ')'
        else:
            name = article
        index2article[i] = name
        index2article_with_references[i] = LegalText(name, translated, index2references[i])
        article2index[name] = i


num_cycles = 0


def DFS_cycles(current_node, visited_nodes, visited_nodes2, visited_nodes3, recursion_stack, recursion_stack2,
               recursion_stack3: list, graph_matrix, index2article_with_references, print_all_chains=False):
    global num_cycles
    current_node_name = index2article_with_references[current_node]
    visited_nodes[current_node] = True
    visited_nodes2[current_node] = current_node_name
    visited_nodes3.append(current_node_name)
    recursion_stack[current_node] = True
    recursion_stack2[current_node] = current_node_name
    recursion_stack3.append(current_node_name)
    # print("Visiting Art.", current_node_name)

    for adjacent_node in range(len(index2article_with_references)):
        adjacent_node_name = index2article_with_references[adjacent_node]
        if graph_matrix[current_node][adjacent_node] == 1:
            # print("Art.", current_node_name, "-> Art.", adjacent_node_name)
            if not visited_nodes[adjacent_node]:
                # print("Art.", current_node_name, "-> Art.", adjacent_node_name)
                DFS_cycles(adjacent_node, visited_nodes, visited_nodes2, visited_nodes3, recursion_stack,
                           recursion_stack2,
                           recursion_stack3, graph_matrix,
                           index2article_with_references, print_all_chains)
            elif recursion_stack[adjacent_node]:
                if graph_matrix[adjacent_node][current_node] == 1:
                    print("Cycle detected:", current_node_name, "<->", adjacent_node_name)
                else:
                    print("Cycle detected: ", end="")
                    for node in range(len(recursion_stack3)):
                        print(recursion_stack3[node], "-> ", end="")
                    print(recursion_stack2[adjacent_node])
                num_cycles += 1

    if print_all_chains and len(recursion_stack3) > 1:
        for node in range(len(recursion_stack3) - 1):
            print(recursion_stack3[node], "-> ", end="")
        print(recursion_stack3[-1])
    recursion_stack[current_node] = False
    recursion_stack2[current_node] = ''
    recursion_stack3.remove(current_node_name)
    # print("Visit on Art.", current_node_name, "completed.")


def detect_cycle(graph_matrix, index2article_with_references):
    num_nodes = len(index2article_with_references)
    global num_cycles
    num_cycles = 0

    visited_nodes = [False] * num_nodes
    visited_nodes2 = [''] * num_nodes
    visited_nodes3 = []
    recursion_stack = [False] * num_nodes
    recursion_stack2 = [''] * num_nodes
    recursion_stack3 = []

    for current_node in range(num_nodes):
        # current_node_name = index2article_with_references[current_node]
        if not visited_nodes[current_node]:
            # print("Running DFS on Art.", current_node_name)
            DFS_cycles(current_node, visited_nodes, visited_nodes2, visited_nodes3, recursion_stack, recursion_stack2,
                       recursion_stack3, graph_matrix,
                       index2article_with_references, print_all_chains=False)


def DFS_kosaraju(graph_matrix, v, visited, visited2, visited3, stack, stack2, index2article_with_references):
    current_node_name = index2article_with_references[v]
    visited[v] = True
    visited2[v] = current_node_name
    visited3.append(current_node_name)
    for adjacent_node in range(len(graph_matrix)):
        if graph_matrix[v][adjacent_node] == 1 and not visited[adjacent_node]:
            adjacent_node_name = index2article_with_references[adjacent_node]
            DFS_kosaraju(graph_matrix, adjacent_node, visited, visited2, visited3, stack, stack2,
                         index2article_with_references)
    stack.append(v)
    stack2.append(current_node_name)


def DFS_util(transposed_graph_matrix, v, visited, visited2, visited3, component, component2,
             index2article_with_references):
    current_node_name = index2article_with_references[v]
    visited[v] = True
    visited2[v] = current_node_name
    visited3.append(current_node_name)
    component.append(v)
    component2.append(current_node_name)
    for adjacent_node in range(len(transposed_graph_matrix)):
        adjacent_node_name = index2article_with_references[adjacent_node]
        if transposed_graph_matrix[v][adjacent_node] == 1 and not visited[adjacent_node]:
            DFS_util(transposed_graph_matrix, adjacent_node, visited, visited2, visited3, component, component2,
                     index2article_with_references)


def kosaraju_scc(graph_matrix, index2article_with_references):
    stack = []
    stack2 = []
    visited = [False] * len(graph_matrix)
    visited2 = [''] * len(graph_matrix)
    visited3 = []

    for i in range(len(graph_matrix)):
        if not visited[i]:
            DFS_kosaraju(graph_matrix, i, visited, visited2, visited3, stack, stack2, index2article_with_references)

    transposed_graph_matrix = graph_matrix.transpose()

    visited = [False] * len(graph_matrix)
    visited2 = [''] * len(graph_matrix)
    visited3 = []
    scc_list = []

    while stack:
        i = stack.pop()
        if not visited[i]:
            component = []
            component2 = []
            DFS_util(transposed_graph_matrix, i, visited, visited2, visited3, component, component2,
                     index2article_with_references)
            scc_list.append(component)

    return scc_list


num_chains = 0


def DFS_chains_of_references(current_node, graph_matrix, index2article_with_references, recursion_stack):
    global num_chains
    current_node_name = index2article_with_references[current_node]
    # print("Visiting Art.", current_node_name)
    recursion_stack.append(current_node_name)

    for adjacent_node in range(len(index2article_with_references)):
        adjacent_node_name = index2article_with_references[adjacent_node]
        if graph_matrix[current_node][adjacent_node] == 1 and adjacent_node_name not in recursion_stack:
            # print("Art.", current_node_name, "-> Art.", adjacent_node_name)
            DFS_chains_of_references(adjacent_node, graph_matrix, index2article_with_references, recursion_stack)

    if len(recursion_stack) > 1:
        num_chains += 1
        for node in range(len(recursion_stack) - 1):
            print(recursion_stack[node], "-> ", end="")
        print(recursion_stack[-1])
    recursion_stack.remove(current_node_name)
    # print("Visit on Art.", current_node_name, "completed.")


def print_chains_of_references(graph_matrix, index2article_with_references):
    global num_chains
    num_chains = 0
    num_nodes = len(index2article_with_references)
    recursion_stack = []

    for current_node in range(num_nodes):
        if index2article_with_references[current_node].translated:
            # current_node_name = index2article_with_references[current_node]
            DFS_chains_of_references(current_node, graph_matrix, index2article_with_references, recursion_stack)


def print_SCC(nodes_list, graph_matrix, index2article):
    G = nx.DiGraph()
    for i in range(len(nodes_list)):
        row = graph_matrix[nodes_list[i]]
        for j in range(len(row)):
            if j in nodes_list and row[j] == 1:
                G.add_edge(nodes_list[i], j)
    G = nx.relabel_nodes(G, index2article)
    nx.draw(G, with_labels=True, node_size=1, font_size=1)
    A = nx.nx_agraph.to_agraph(G)
    A.node_attr.update(shape="box", color="black", fontname="Helvetica",
                       fontsize=12, fontcolor="black", fonttype="bold", nodesep=2.0,
                       width=0.65, height=0.3, margin=0)
    A.edge_attr.update(color="blue", style="solid", penwidth=0.5, arrowsize=0.6, arrowhead="vee")
    # print(A)
    A.layout(prog="sfdp", args="-Goverlap=prism")
    # A.draw("SCC_with_59_nodes_option2.png")


def main():
    option = 0
    index2article: dict[int, str] = {}
    index2article_with_references: dict[int, LegalText] = {}
    article2index: dict[str, int] = {}
    df = pd.read_excel('GDPR_map.xlsx',
                       names=['Article', 'Paragraph', 'Point', 'Translated', 'References', 'Link'],
                       dtype={'Article': str, 'Paragraph': str, 'Point': str, 'Translated': bool, 'References': str,
                              'Link': str})
    df['Link'] = setup_links()
    df.fillna('', inplace=True)
    # print(df)
    setup_dictionaries(df, index2article, index2article_with_references, article2index)
    # print(index2article)
    # print(index2article_with_references)# maps the number node to correct name
    # print(article2index)

    # choose an option between 0, 1, 2, 3 to set up the graph adjacency matrix
    graph_matrix = setup_graph_matrix(index2article_with_references, article2index, option)

    # with np.printoptions(threshold=np.inf):
    # print(graph_matrix)

    detect_cycle(graph_matrix, index2article_with_references)
    if num_cycles > 0:
        print(f"{num_cycles} cycle(s) detected.\n")
    else:
        print("No cycle detected.\n")

    scc_list = kosaraju_scc(graph_matrix, index2article_with_references)
    for i in range(len(scc_list)):
        if len(scc_list[i]) > 1:
            print(f"SCC: {[index2article_with_references[node].name for node in scc_list[i]]}")
            print("Number of nodes in the strongly connected component:", len(scc_list[i]))
            if i == 217:
                print_SCC(scc_list[i], graph_matrix, index2article_with_references)
    print("Number of non-singleton strongly connected components:", len([scc for scc in scc_list if len(scc) > 1]))
    print("Total number of strongly connected components:", len(scc_list))
    """
    if option < 2:
        print("\nPrinting the chains of references for the translated articles:")
        print_chains_of_references(graph_matrix, index2article_with_references)
        print("Number of chains of references:", num_chains)
    """
    # Author: Esteban Garcia Taquez

    # turn the graph adjacency matrix into 2d array
    arrayMatrix = np.array(graph_matrix)

    # create the graph from the matrix as an array
    G = nx.DiGraph(arrayMatrix)

    # modify the names of the nodes
    G = nx.relabel_nodes(G, index2article)

    # remove nodes with degree lower than 1
    # print(G)
    degree = G.degree()
    remove = []
    for i in degree:
        if i[1] < 1:
            remove.append(i[0])
    G.remove_nodes_from(remove)
    # print(G)

    # draw and print the graph
    nx.draw(G, with_labels=True, node_size=1, font_size=1)
    A = nx.nx_agraph.to_agraph(G)
    A.node_attr.update(shape="box", color="black", fontname="Helvetica",
                       fontsize=12, fontcolor="black", fonttype="bold", nodesep=2.0,
                       width=0.65, height=0.3, margin=0)
    A.edge_attr.update(color="blue", style="solid", penwidth=0.5, arrowsize=0.6, arrowhead="vee")
    # print(A)
    A.layout(prog="sfdp", args="-Goverlap=prism")
    A.draw("graph_option0_3.jpg")


if __name__ == '__main__':
    main()
